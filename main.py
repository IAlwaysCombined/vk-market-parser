import os
import sys
from time import sleep
from typing import List, Dict

import openpyxl
import requests
from dotenv import load_dotenv

import category_map as cm

load_dotenv()

MAX_IMAGE_SIZE_MB = 6  # Максимальный размер файла
ITEMS_PER_REQUEST = 200  # Максимальное количество товаров за один запрос
REQUEST_DELAY = 0.5  # Задержка между запросами в секундах

output_dir = 'output'
OUTPUT_FILENAME = 'vk_market_data.xlsx'


def map_category(category):
    """Маппинг категорий вк на наши категории товаров"""
    return cm.mapping.get(category, category)


def print_progress(current, total):
    """Выводит прогресс-бар в консоль"""
    progress = current / total
    bar_length = 40
    filled_length = int(bar_length * progress)
    bar = '█' * filled_length + '-' * (bar_length - filled_length)
    sys.stdout.write(f'\rПрогресс: [{bar}] {current}/{total} ({progress:.1%})')
    sys.stdout.flush()


def get_env_var(name, default=None):
    """Получает переменную окружения с возможным значением по умолчанию"""
    value = os.getenv(name)
    return value if value is not None else default


def get_image_url(photos, index):
    """Получение фотографий с определенным размером"""
    if not photos or index >= len(photos):
        return None

    photo = photos[index]
    photo_sizes = sorted(photo["sizes"], key=lambda x: x["height"] * x["width"], reverse=True)

    for size in photo_sizes:
        image_url = size["url"]
        try:
            image_size_mb = len(requests.get(image_url).content) / (1024 * 1024)
            if image_size_mb <= MAX_IMAGE_SIZE_MB:
                return image_url
        except requests.exceptions.RequestException:
            continue

    return None


def get_primary_image_url(item):
    """Получение 1 фотографии"""
    if "photos" not in item:
        return None
    return get_image_url(item["photos"], 0)


def get_secondary_image_url(item):
    """Получение 2 фотографии"""
    if "photos" not in item or len(item["photos"]) < 2:
        return None
    return get_image_url(item["photos"], 1)


def create_excel_file(file_name: str, items: List[Dict]) -> None:
    """
    Создает Excel файл с данными товаров.

    Args:
        file_name: Имя файла (без расширения)
        items: Отсортированный список товаров
    """
    print("\nСоздание Excel файла...")

    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Товары VK Market"

        # Заголовки столбцов
        headers = [
            "ID", "Категория", "Название",
            "Цена", "Описание",
            "Основное изображение", "Дополнительное изображение"
        ]
        for col_num, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_num, value=header)

        # Заполнение данных
        for row_num, item in enumerate(items, start=2):
            worksheet.cell(row=row_num, column=1, value=item["id"])
            worksheet.cell(row=row_num, column=2, value=map_category(item["category"]["name"]))
            worksheet.cell(row=row_num, column=3, value=item["title"])

            # Обработка цены
            price = item.get("price", {}).get("amount", "")
            worksheet.cell(row=row_num, column=4, value=price[:-2] if price else "")

            worksheet.cell(row=row_num, column=5, value=item["description"])
            worksheet.cell(row=row_num, column=6, value=get_primary_image_url(item))
            worksheet.cell(row=row_num, column=7, value=get_secondary_image_url(item))

        # Сохранение файла
        os.makedirs("output", exist_ok=True)
        file_path = f"output/{file_name}.xlsx"
        workbook.save(file_path)

        print(f"Файл успешно создан: {file_path}")
        print(f"Экспортировано товаров: {len(items)}")

    except Exception as e:
        print(f"Ошибка при создании файла: {e}")
        raise


def get_all_vk_market_items(access_token: str, owner_id: str) -> List[Dict]:
    """
    Получает все товары из VK Market.

    Args:
        access_token: Токен доступа VK API
        owner_id: ID владельца товаров

    Returns:
        Список товаров, отсортированный по ID
    """
    print("Загрузка товаров из VK Market...")

    all_items = []
    offset = 0
    total_items = None
    vk_version = "5.131"
    vk_market_url = get_env_var('VK_MARKET_URL')
    items_per_request = 100
    request_delay = 0.5

    while True:
        params = {
            "owner_id": owner_id,
            "count": items_per_request,
            "offset": offset,
            "access_token": access_token,
            "v": vk_version,
            "extended": 1
        }

        try:
            response = requests.get(vk_market_url, params=params)
            data = response.json()

            if "error" in data:
                print(f"Ошибка API: {data['error']['error_msg']}")
                break

            if total_items is None:
                total_items = data["response"]["count"]
                print(f"Всего товаров: {total_items}")

            items = data["response"]["items"]
            if not items:
                break

            all_items.extend(items)
            offset += len(items)

            print(f"Загружено: {len(all_items)}/{total_items} товаров")

            if len(all_items) >= total_items:
                break

            sleep(request_delay)

        except Exception as e:
            print(f"Ошибка при запросе: {e}")
            break

    # Сортировка по ID
    all_items.sort(key=lambda x: x["id"])
    print("Товары отсортированы по ID")

    return all_items


def main():
    """Основная функция выполнения программы."""
    print("=== Экспорт товаров VK Market ===")

    # Получение учетных данных
    access_token = os.getenv("ACCESS_TOKEN")
    owner_id = os.getenv("OWNER_ID")

    if not all([access_token, owner_id]):
        print("Ошибка: Необходимо указать ACCESS_TOKEN и OWNER_ID в переменных окружения")
        return

    try:
        # Получение и сортировка товаров
        items = get_all_vk_market_items(access_token, owner_id)

        if not items:
            print("Нет товаров для экспорта")
            return

        # Экспорт в Excel
        create_excel_file("vk_market_export", items)

        print("Экспорт завершен успешно!")

    except Exception as e:
        print(f"Критическая ошибка: {e}")


if __name__ == "__main__":
    main()
